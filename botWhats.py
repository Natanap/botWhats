from time import sleep

import openpyxl
import re
import pandas as pd
import datetime
import os
import shutil

from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

class WhatsAppBot:
    

    def __init__(self) -> None: 
        print("Estamos iniciado a varredura por favor aguarde...")
        sleep(5)
        print("Devido a sincronização da plataforma Whatsapp, a busca poderá demorar um pouco...\n***Atenção: Caso não localize algum pedido recente reinicie o programa novamente.***")
        user_data_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Google", "Chrome", "User Data", "Default")
    
        self.options = webdriver.ChromeOptions()
        self.options.add_argument(f'user-data-dir={user_data_dir}')

        # Remove unnecessary logs
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])

        self.__selenium_driver()

        self.__open_whatsapp()

        self.__search_orders()

        self.__get_all_chats()



    def __open_whatsapp(self) -> None:
        self.driver.get(url="https://web.whatsapp.com")
        sleep(60)

    def __selenium_driver(self) -> None: 
        service = Service(ChromeDriverManager().install())   
        self.driver = webdriver.Chrome(service=service, options=self.options)
        self.driver.implicitly_wait(30)

    def __search_orders(self) -> None:
        sleep(10)
        element = self.driver.find_element(By.XPATH, '//*[@id="side"]/div[1]/div/div/div[2]/div/div[1]/p')
        element.click()
        element.send_keys("Total: R$")
        sleep(5)

    def __get_all_chats(self) -> None:
        elements = self.driver.find_elements(By.XPATH, '//*[@id="pane-side"]/div[1]/div')
        values = []
        total_pedidos = 0.0
        for element in elements:
            # Obtem o texto completo do elemento div
            texto_completo = element.get_attribute("textContent").strip()
            result = texto_completo.replace("\n", " ")
            padrao = "\d{2}:\d{2}"
            horas = re.findall(padrao, result)

            for hora in horas:
                index_hora = result.find(hora)
                while index_hora >= 0:
                    texto_pagamento = ""
                    valores = 0.0
                    index_pagamento = -1
                    for palavra in ["Dinheiro", "Cartão", "Pix"]:
                        index_pagamento_palavra = result.lower().find(palavra.lower(), index_hora+len(hora))
                        if index_pagamento_palavra >= 0 and (index_pagamento == -1 or index_pagamento_palavra < index_pagamento):
                            index_pagamento = index_pagamento_palavra
                            texto_pagamento = palavra
                    if index_pagamento >= 0:
                        index_total = result.find("Total: R$ ", index_pagamento)
                        if index_total >= 0:
                            index_valor = index_total + len("Total: R$ ")
                            valor_str = ""
                            while index_valor < len(result) and (result[index_valor].isdigit() or result[index_valor] == ","):
                                if result[index_valor] == ',':
                                    valor_str += "."
                                else:
                                    valor_str += result[index_valor]
                                index_valor += 1
                            valores = float(valor_str)
                            total_pedidos += valores
                        values.append([hora, texto_pagamento.capitalize(), valores])
                    index_hora = result.find(hora, index_hora+len(hora))
            self.salvar_pedidos_excel(values, total_pedidos)
            
    

    def salvar_pedidos_excel(self, values, total_pedidos):
        pasta_destino='Fluxo_Caixa'
        # Define o caminho completo da pasta
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        pasta_completa = os.path.join(desktop_path, pasta_destino)

        # Verifica se a pasta já existe, caso contrário, cria a pasta
        if not os.path.exists(pasta_completa):
            os.makedirs(pasta_completa)

        # Obter data atual
        data = datetime.date.today().strftime('%Y-%m-%d')
        # Nome do arquivo
        file_name = f'pedidos-do-dia({data}).xlsx'
        file_path = os.path.join(pasta_completa, file_name)

        # Verificar se o arquivo já existe, adicionar número ao final se necessário
        if os.path.exists(file_path):
            i = 1
            while os.path.exists(f"{file_path.split('.')[0]}_{i}.xlsx"):
                i += 1
            file_name = f"{file_name.split('.')[0]}_{i}.xlsx"
            file_path = os.path.join(pasta_completa, file_name)

        # Criar novo arquivo ou abrir arquivo existente
        wb = openpyxl.Workbook()
        # Selecionar planilha ativa
        sheet = wb.active
        # Adicionar títulos às colunas
        sheet.cell(row=1, column=1).value = "Hora"
        sheet.cell(row=1, column=2).value = "Tipo de Pagamento"
        sheet.cell(row=1, column=3).value = "Valores"
        # Escrever valores na planilha
        if values:
            for i, valor in enumerate(values):
                # Adicionar os valores nas células da planilha, incrementando o número da linha a cada novo valor encontrado
                sheet.cell(row=i+2, column=1).value = valor[0]
                sheet.cell(row=i+2, column=2).value = valor[1]
                sheet.cell(row=i+2, column=3).value = valor[2]
            # Adicionar o total de pedidos na última linha
            sheet.cell(row=len(values)+2, column=1).value = "Total"
            sheet.cell(row=len(values)+2, column=2).value = ""
            sheet.cell(row=len(values)+2, column=3).value = total_pedidos
            # Salvar arquivo
            wb.save(file_path)
            print(f'Valores encontrados salvos em na sua Área de trabalho pasta Fluxo Caixa"{file_name}".')
            sleep(5)
        else:
            print('Nenhum valor encontrado.')

        

if __name__ == "__main__":
    WhatsAppBot()